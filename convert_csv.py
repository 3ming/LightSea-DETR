import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import argparse

def clean_csv_and_convert_to_excel(input_file_path, output_excel_path):
    # 读取 CSV 文件
    df = pd.read_csv(input_file_path)

    # 清理列名
    df.columns = [header.strip() for header in df.columns]

    # 创建包含 "min" 和 "max" 的行
    min_max_row = [''] * len(df.columns)
    for col in [2, 3, 4, 9, 10, 11]:
        min_max_row[col - 1] = 'min'  # 列索引从1开始，所以减1
    for col in [5, 6, 7, 8]:
        min_max_row[col - 1] = 'max'

    # 在第一行下新增第二行和第三行的空白行
    blank_row = pd.DataFrame([[''] * len(df.columns)], columns=df.columns)
    min_max_row_df = pd.DataFrame([min_max_row], columns=df.columns)
    df = pd.concat([df.iloc[:0], min_max_row_df, blank_row, df], ignore_index=True)

    # 计算特定列的最小值和最大值，并将其放入第三行
    min_columns = [1, 2, 3, 8, 9, 10]  # 列索引调整为从0开始
    max_columns = [4, 5, 6, 7]
    for col in min_columns:
        df.iloc[1, col] = df.iloc[2:, col].min()
    for col in max_columns:
        df.iloc[1, col] = df.iloc[2:, col].max()

    # 使用 ExcelWriter 保存为 Excel 文件
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=True, sheet_name='Sheet1')

        # 获取 openpyxl 对象
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # 冻结窗口，冻结第1行到第3行
        worksheet.freeze_panes = 'A4'

        # 设置字体加粗
        font = Font(bold=True)
        for col in range(1, df.shape[1] + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = font

        # 设置列宽自适应
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        # 设置第二行和第三行的内容为右对齐
        for row in [2, 3]:
            for cell in worksheet[row]:
                cell.alignment = Alignment(horizontal='right')


def main():
    # 设置命令行参数解析
    parser = argparse.ArgumentParser(description='Convert CSV file to Excel and perform data manipulation.')
    parser.add_argument('--name', '-n', type=str, required=True, help='Name to use in file paths')

    # 解析命令行参数
    args = parser.parse_args()
    result_name = args.name

    # 设置文件路径
    input_file_path = f'runs/train/{result_name}/results.csv'
    output_excel_path = f'runs/train/{result_name}/results.xlsx'

    # 执行转换函数
    clean_csv_and_convert_to_excel(input_file_path, output_excel_path)

if __name__ == "__main__":
    main()
